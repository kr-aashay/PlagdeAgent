# ═══════════════════════════════════════════════════════════════════════════
# AI Ethics Pledge — Database Seeding Python Script (Excel version)
# ═══════════════════════════════════════════════════════════════════════════

import os
import time
import openpyxl
import psycopg2
from psycopg2.extras import execute_values
from dotenv import load_dotenv

def main():
    load_dotenv()
    
    DATABASE_URL = os.getenv("DATABASE_URL")
    if not DATABASE_URL:
        print("❌ ERROR: DATABASE_URL not found in .env")
        exit(1)
        
    t_start = time.time()
    
    # 1. Connect to PostgreSQL
    print("🔗 Connecting to PostgreSQL...")
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()
    
    # 2. Clean existing tables
    print("🧹 Truncating students and employees tables...")
    cur.execute("TRUNCATE TABLE students, employees RESTART IDENTITY CASCADE;")
    conn.commit()
    print("✅ Tables cleaned.\n")
    
    # 3. Read students onroll Excel
    print("🎓 Reading students_onroll_22082026.xlsx...")
    t0 = time.time()
    wb_students = openpyxl.load_workbook("students_onroll_22082026.xlsx", read_only=True, data_only=True)
    sheet_students = wb_students.active
    
    students_data = []
    seen_reg = set()
    
    # Column indices: 0=registerno, 1=name, 2=vuid, 3=coursename, 
    #                 4=branch_shortname, 5=branchname, 6=cyear, 7=sectioncode
    for row in sheet_students.iter_rows(min_row=2, values_only=True):
        registerno = row[0]
        name = row[1]
        vuid = row[2]
        coursename = row[3]
        branch_shortname = row[4]
        branchname = row[5]
        cyear = row[6]
        sectioncode = row[7]
        
        if registerno and name:
            registerno_str = str(registerno).strip().upper()
            name_str = str(name).strip()
            vuid_str = str(vuid).strip() if vuid else ""
            coursename_str = str(coursename).strip() if coursename else ""
            branch_shortname_str = str(branch_shortname).strip() if branch_shortname else ""
            branchname_str = str(branchname).strip() if branchname else ""
            cyear_str = str(cyear).strip() if cyear else ""
            sectioncode_str = str(sectioncode).strip() if sectioncode else ""
            
            if registerno_str not in seen_reg:
                seen_reg.add(registerno_str)
                students_data.append((
                    registerno_str, name_str, vuid_str, coursename_str,
                    branch_shortname_str, branchname_str, cyear_str, sectioncode_str
                ))
                
    print(f"📖 Parsed {len(students_data)} unique students in {time.time() - t0:.2f} seconds.\n")
    
    # 4. Read employees onroll Excel
    print("💼 Reading employee_onroll_22082026.xlsx...")
    t0 = time.time()
    wb_employees = openpyxl.load_workbook("employee_onroll_22082026.xlsx", read_only=True, data_only=True)
    sheet_employees = wb_employees.active
    
    employees_data = []
    seen_emp = set()
    
    # Column indices:
    # 0: empcode, 1: name, 3: deptname
    for row in sheet_employees.iter_rows(min_row=2, values_only=True):
        code = row[0]
        name = row[1]
        dept = row[3] # deptname is 4th column (index 3)
        
        if code and name:
            code_str = str(code).strip().upper()
            name_str = str(name).strip()
            dept_str = str(dept).strip() if dept else ""
            
            if code_str not in seen_emp:
                seen_emp.add(code_str)
                employees_data.append((name_str, code_str, dept_str))
                
    print(f"📖 Parsed {len(employees_data)} unique employees in {time.time() - t0:.2f} seconds.\n")
    
    # 5. Bulk insert students
    print("⚙️  Bulk inserting students to database...")
    t0 = time.time()
    execute_values(
        cur,
        """INSERT INTO students 
           (registerno, name, vuid, coursename, branch_shortname, branchname, cyear, sectioncode) 
           VALUES %s""",
        students_data
    )
    conn.commit()
    print(f"✅ Seeding students complete in {time.time() - t0:.2f} seconds.")
    
    # 6. Bulk insert employees
    print("⚙️  Bulk inserting employees to database...")
    t0 = time.time()
    execute_values(
        cur,
        "INSERT INTO employees (name, employee_id, department) VALUES %s",
        employees_data
    )
    conn.commit()
    print(f"✅ Seeding employees complete in {time.time() - t0:.2f} seconds.")
    
    cur.close()
    conn.close()
    print(f"\n🎉 Seeding finished successfully in {time.time() - t_start:.2f} seconds total!")

if __name__ == "__main__":
    main()
